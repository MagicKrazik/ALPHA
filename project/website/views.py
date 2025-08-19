# website/views.py - Fixed imports at the top
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model, login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.db.models import Count, Avg, Q, Sum, F, Case, When, IntegerField
from django.db.models.functions import TruncMonth, TruncWeek, ExtractHour
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.http import JsonResponse, FileResponse, HttpResponse
from django.utils import timezone
from django.utils.encoding import force_str
from django.utils.http import urlsafe_base64_decode
from django.utils.translation import gettext_lazy as _
from django.conf import settings

import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Local imports
from .models import Patient, PreSurgeryForm, PostDuringSurgeryForm, MedicoUser
from .forms import (
    MedicRegistrationForm, ContactForm, PatientForm, 
    PreSurgeryCreateForm, PostSurgeryCreateForm
)

# Try to import risk assessment, create fallback if not available
try:
    from .utils.risk_assessment import RiskCalculator, AdvancedRiskCalculator
except ImportError:
    # Fallback risk calculator if the advanced one fails
    class RiskCalculator:
        @staticmethod
        def calculate_airway_risk(form):
            # Basic fallback calculation
            risk_score = 0
            risk_factors = []
            
            if hasattr(form, 'mallampati') and form.mallampati >= 3:
                risk_score += 25
                risk_factors.append(f"Mallampati {form.mallampati}")
            
            if hasattr(form, 'estado_fisico_asa') and form.estado_fisico_asa >= 4:
                risk_score += 30
                risk_factors.append(f"ASA {form.estado_fisico_asa}")
            
            return min(risk_score, 100), risk_factors
        
        @staticmethod
        def get_risk_level(score):
            if score >= 70:
                return {
                    'level': 'ALTO',
                    'color': '#e74c3c',
                    'description': 'Alto riesgo de vía aérea difícil',
                    'recommendations': ['Preparación especial requerida']
                }
            elif score >= 40:
                return {
                    'level': 'MODERADO',
                    'color': '#f39c12',
                    'description': 'Riesgo moderado',
                    'recommendations': ['Monitoreo adicional']
                }
            else:
                return {
                    'level': 'BAJO',
                    'color': '#27ae60',
                    'description': 'Riesgo bajo',
                    'recommendations': ['Manejo estándar']
                }


# Your existing view functions start here...
User = get_user_model()

def home(request):
    return render(request, 'home.html')

def register(request):
    if request.method == 'POST':
        form = MedicRegistrationForm(request.POST)
        if form.is_valid():
            try:
                # Check if email already exists
                email = form.cleaned_data['email']
                if User.objects.filter(email=email).exists():
                    messages.error(request, 'Este correo electrónico ya está registrado.')
                    return render(request, 'register.html', {'form': form})

                # Create active user directly
                user = User(
                    username=form.cleaned_data['usuario'],
                    email=email,
                    nombre=form.cleaned_data['nombre'],
                    apellidos=form.cleaned_data['apellidos'],
                    telefono=form.cleaned_data['telefono'],
                    especialidad=form.cleaned_data['especialidad'][:3].upper(),
                    is_active=True  # Set user as active immediately
                )
                user.set_password(form.cleaned_data['password1'])
                user.save()
                
                # Log the user in automatically
                login(request, user)
                messages.success(request, '¡Registro exitoso! Bienvenido.')
                return redirect('home')
                
            except IntegrityError as e:
                if 'UNIQUE constraint' in str(e) and 'username' in str(e):
                    messages.error(request, 'Este nombre de usuario ya está en uso.')
                elif 'UNIQUE constraint' in str(e) and 'email' in str(e):
                    messages.error(request, 'Este correo electrónico ya está registrado.')
                else:
                    messages.error(request, 'Error al crear el usuario. Por favor, intente nuevamente.')
            except Exception as e:
                messages.error(request, 'Error al crear el usuario. Por favor, intente nuevamente.')
    else:
        form = MedicRegistrationForm()
    
    return render(request, 'register.html', {'form': form})


# login and logout view:

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, 'Inicio de sesión exitoso.')
            return redirect('home')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('home')

# contact form view:

def contact(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            # Save the message
            contact_message = form.save()
            
            # Send email notification
            subject = f"Nuevo mensaje de contacto - {form.cleaned_data['subject']}"
            message = f"""
            Nuevo mensaje de contacto recibido:
            
            Nombre: {form.cleaned_data['name']}
            Email: {form.cleaned_data['email']}
            Teléfono: {form.cleaned_data['phone']}
            Asunto: {form.cleaned_data['subject']}
            Mensaje: {form.cleaned_data['message']}
            """
            
            try:
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [settings.CONTACT_EMAIL],
                    fail_silently=False,
                )
                messages.success(request, 'Mensaje enviado correctamente. Nos pondremos en contacto pronto.')
            except Exception as e:
                messages.warning(request, 'El mensaje se ha guardado pero hubo un problema al enviar la notificación.')
            
            return redirect('contact')
    else:
        form = ContactForm()
    
    context = {
        'form': form,
        'company_email': 'contacto@alphaproject.com',  # Dummy email
        'company_phone': '+52 (555) 123-4567',  # Dummy phone
        'company_address': 'Ciudad de México, México'  # Dummy address
    }
    
    return render(request, 'contact.html', context)


## patient registration views:

@login_required
def patient_list(request):
    """
    List view for patients with search and form type filtering
    """
    search_query = request.GET.get('q', '')
    form_type = request.GET.get('form', '')
    patients = Patient.objects.filter(medico=request.user, activo=True)
    
    if search_query:
        patients = patients.filter(
            Q(folio_hospitalizacion__icontains=search_query) |
            Q(nombres__icontains=search_query) |
            Q(apellidos__icontains=search_query)
        )
    
    context = {
        'patients': patients,
        'search_query': search_query,
        'form_type': form_type  # Pass the form type to the template
    }
    
    # If a specific form type is requested, use the appropriate template
    if form_type == 'pre':
        return render(request, 'presurgery_list.html', context)
    elif form_type == 'post':
        return render(request, 'postsurgery_list.html', context)
    
    return render(request, 'patient_list.html', context)

@login_required
def patient_create(request):
    """
    Create view for new patient registration
    """
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.medico = request.user
            patient.save()
            messages.success(request, _('Paciente registrado exitosamente.'))
            return redirect('patient-detail', patient.id_paciente)
    else:
        form = PatientForm()
    
    return render(request, 'patient_form.html', {'form': form})

@login_required
def patient_update(request, pk):
    """
    Update view for existing patient information
    """
    patient = get_object_or_404(Patient, id_paciente=pk, medico=request.user)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, _('Información del paciente actualizada.'))
            return redirect('patient-detail', pk)
    else:
        form = PatientForm(instance=patient)
    
    return render(request, 'patient_form.html', {
        'form': form,
        'patient': patient
    })

@login_required
def patient_detail(request, pk):
    """
    Detail view for patient information
    """
    patient = get_object_or_404(Patient, id_paciente=pk)
    
    # Check if user has permission to view this patient
    if patient.medico != request.user and not request.user.is_staff:
        messages.error(request, _('No tienes permiso para ver esta información.'))
        return redirect('patient-list')
    
    # Get pre-surgery forms for this patient
    presurgery_forms = PreSurgeryForm.objects.filter(
        folio_hospitalizacion__startswith=f"PRE-{patient.folio_hospitalizacion}"
    ).order_by('-fecha_reporte')
    
    # Get post-surgery forms for this patient
    postsurgery_forms = PostDuringSurgeryForm.objects.filter(
        folio_hospitalizacion__folio_hospitalizacion__in=presurgery_forms.values_list('folio_hospitalizacion', flat=True)
    ).order_by('-folio_hospitalizacion__fecha_reporte')
    
    context = {
        'patient': patient,
        'presurgery_forms': presurgery_forms,
        'postsurgery_forms': postsurgery_forms,
    }
    
    return render(request, 'patient_detail.html', context)

@login_required
def patient_delete(request, pk):
    """
    Delete view for patient (soft delete)
    """
    patient = get_object_or_404(Patient, id_paciente=pk, medico=request.user)
    
    if request.method == 'POST':
        patient.activo = False
        patient.save()
        messages.success(request, _('Paciente eliminado exitosamente.'))
        return redirect('patient-list')
    
    return render(request, 'patient_confirm_delete.html', {
        'patient': patient
    })


## pre and post surgery forms:

@login_required
def presurgery_create(request, patient_id):
    patient = get_object_or_404(Patient, id_paciente=patient_id, medico=request.user)
    
    try:
        # Check if a form already exists for this patient
        existing_form = PreSurgeryForm.objects.filter(
            folio_hospitalizacion=f"PRE-{patient.folio_hospitalizacion}"
        ).first()
        
        if existing_form:
            messages.warning(request, 'Ya existe un formulario pre-quirúrgico para este paciente.')
            return redirect('presurgery-detail', pk=existing_form.folio_hospitalizacion)

        if request.method == 'POST':
            form = PreSurgeryCreateForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    presurgery = form.save(commit=False)
                    presurgery.folio_hospitalizacion = f"PRE-{patient.folio_hospitalizacion}"
                    presurgery.fecha_reporte = timezone.now().date()
                    presurgery.medico = request.user.get_full_name()
                    presurgery.save()
                    
                    messages.success(request, 'Formulario pre-quirúrgico creado exitosamente.')
                    return redirect('patient-detail', patient_id)
                except Exception as e:
                    messages.error(request, f'Error al guardar el formulario: {str(e)}')
            else:
                error_messages = []
                for field, errors in form.errors.items():
                    error_messages.append(f"{field}: {', '.join(errors)}")
                messages.error(request, f'Por favor corrija los siguientes errores: {"; ".join(error_messages)}')
        else:
            initial_data = {
                'nombres': patient.nombres,
                'apellidos': patient.apellidos,
                'fecha_nacimiento': patient.fecha_nacimiento,
                'medico': request.user.get_full_name(),
                'fecha_reporte': timezone.now().date(),
                'medico_tratante': request.user.get_full_name(),
            }
            form = PreSurgeryCreateForm(initial=initial_data)

        return render(request, 'presurgery_form.html', {
            'form': form,
            'patient': patient
        })

    except Exception as e:
        messages.error(request, f'Error inesperado: {str(e)}')
        return redirect('patient-list')

@login_required
def presurgery_detail(request, pk):
    presurgery = get_object_or_404(PreSurgeryForm, folio_hospitalizacion=pk)
    patient = get_object_or_404(Patient, folio_hospitalizacion=pk.replace('PRE-', ''))
    
    # Check if user has permission to view this form
    if patient.medico != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para ver este formulario.')
        return redirect('patient-list')
    
    return render(request, 'presurgery_detail.html', {
        'form': presurgery,
        'patient': patient
    })

@login_required
def presurgery_update(request, pk):
    presurgery = get_object_or_404(PreSurgeryForm, folio_hospitalizacion=pk)
    patient = get_object_or_404(Patient, folio_hospitalizacion=pk.replace('PRE-', ''))
    
    if patient.medico != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para editar este formulario.')
        return redirect('patient-list')
    
    if request.method == 'POST':
        form = PreSurgeryCreateForm(request.POST, request.FILES, instance=presurgery)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Formulario actualizado exitosamente.')
                return redirect('presurgery-detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar el formulario: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = PreSurgeryCreateForm(instance=presurgery)
    
    return render(request, 'presurgery_form.html', {
        'form': form,
        'presurgery': presurgery,
        'patient': patient
    })


# In views.py - Replace the postsurgery_create function
@login_required
def postsurgery_create(request, patient_id):
    """Create post-surgery form - FIXED VERSION"""
    print(f"DEBUG: Starting postsurgery_create for patient {patient_id}")
    print(f"DEBUG: Request method: {request.method}")
    
    if request.method == 'POST':
        print(f"DEBUG: POST data: {request.POST}")
        print(f"DEBUG: FILES data: {request.FILES}")
    patient = get_object_or_404(Patient, id_paciente=patient_id, medico=request.user)
    
    # Get the pre-surgery form
    try:
        presurgery = PreSurgeryForm.objects.get(
            folio_hospitalizacion=f"PRE-{patient.folio_hospitalizacion}"
        )
    except PreSurgeryForm.DoesNotExist:
        messages.error(request, 'Debe completar el formulario pre-quirúrgico antes de crear el post-quirúrgico.')
        return redirect('patient-detail', patient_id)
    
    # Check if post-surgery form already exists
    try:
        existing_post = PostDuringSurgeryForm.objects.get(folio_hospitalizacion=presurgery)
        messages.warning(request, 'Ya existe un formulario post-quirúrgico para este paciente.')
        return redirect('postsurgery-detail', pk=presurgery.folio_hospitalizacion)
    except PostDuringSurgeryForm.DoesNotExist:
        pass  # This is what we want - no existing form

    if request.method == 'POST':
        form = PostSurgeryCreateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create the instance but don't save yet
                postsurgery = form.save(commit=False)
                postsurgery.folio_hospitalizacion = presurgery
                
                # Set default values if not provided
                if not postsurgery.nombre_anestesiologo:
                    postsurgery.nombre_anestesiologo = request.user.get_full_name()
                
                if not postsurgery.especialidad:
                    postsurgery.especialidad = request.user.get_especialidad_display()
                
                # Save the instance
                postsurgery.save()
                
                messages.success(request, 'Formulario post-quirúrgico creado exitosamente.')
                return redirect('patient-detail', patient_id)
                
            except ValidationError as e:
                # Handle validation errors
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
            except Exception as e:
                messages.error(request, f'Error al guardar el formulario: {str(e)}')
        else:
            # Handle form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        # GET request - create empty form with initial data
        initial_data = {
            'nombre_anestesiologo': request.user.get_full_name(),
            'especialidad': request.user.get_especialidad_display() if hasattr(request.user, 'get_especialidad_display') else '',
        }
        form = PostSurgeryCreateForm(initial=initial_data)

    return render(request, 'postsurgery_form.html', {
        'form': form,
        'patient': patient,
        'presurgery': presurgery
    })

@login_required
def postsurgery_detail(request, pk):
    """
    View for displaying post-surgery form details
    """
    postsurgery = get_object_or_404(PostDuringSurgeryForm, folio_hospitalizacion__folio_hospitalizacion=pk)
    # Get the related patient through the pre-surgery form
    patient = Patient.objects.get(folio_hospitalizacion=pk.replace('PRE-', ''))
    
    # Check permissions
    if patient.medico != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para ver este formulario.')
        return redirect('patient-list')
    
    return render(request, 'postsurgery_detail.html', {
        'form': postsurgery,
        'patient': patient
    })

@login_required
def postsurgery_update(request, pk):
    """
    Update view for post-surgery form
    """
    postsurgery = get_object_or_404(PostDuringSurgeryForm, folio_hospitalizacion__folio_hospitalizacion=pk)
    patient = Patient.objects.get(folio_hospitalizacion=pk.replace('PRE-', ''))
    
    # Check permissions
    if patient.medico != request.user and not request.user.is_staff:
        messages.error(request, _('No tienes permiso para editar este formulario.'))
        return redirect('patient-detail', pk=patient.id_paciente)
    
    if request.method == 'POST':
        form = PostSurgeryCreateForm(request.POST, request.FILES, instance=postsurgery)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, _('Formulario post-quirúrgico actualizado exitosamente.'))
                return redirect('postsurgery-detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar el formulario: {str(e)}')
        else:
            messages.error(request, _('Por favor corrija los errores en el formulario.'))
    else:
        form = PostSurgeryCreateForm(instance=postsurgery)
    
    return render(request, 'postsurgery_form.html', {
        'form': form,
        'patient': patient,
        'postsurgery': postsurgery
    })


## Dashboard view:

@login_required
def dashboard(request):
    """Enhanced dashboard with AI risk assessment and real-time alerts"""
    
    # Get current user's patients and forms
    user_patients = Patient.objects.filter(medico=request.user, activo=True)
    
    # Pre-surgery forms for this user
    presurgery_forms = PreSurgeryForm.objects.filter(
        medico=request.user.get_full_name()
    ).select_related().prefetch_related('post_surgery_form')
    
    # Post-surgery forms
    postsurgery_forms = PostDuringSurgeryForm.objects.filter(
        folio_hospitalizacion__in=presurgery_forms.values('folio_hospitalizacion')
    )
    
    # Basic Statistics
    total_patients = user_patients.count()
    total_surgeries = presurgery_forms.count()
    completed_surgeries = postsurgery_forms.count()
    pending_surgeries = total_surgeries - completed_surgeries
    
    patients_this_month = user_patients.filter(
        fecha_registro__month=datetime.now().month,
        fecha_registro__year=datetime.now().year
    ).count()
    
    # Risk Assessment Analysis
    risk_assessments = []
    high_risk_alerts = []
    critical_patients = []
    
    for form in presurgery_forms:
        try:
            risk_score, risk_factors = RiskCalculator.calculate_airway_risk(form)
            risk_level = RiskCalculator.get_risk_level(risk_score)
            
            assessment = {
                'form': form,
                'patient_name': f"{form.nombres} {form.apellidos}",
                'folio': form.folio_hospitalizacion,
                'risk_score': risk_score,
                'risk_level': risk_level['level'],
                'risk_color': risk_level['color'],
                'risk_factors': risk_factors,
                'recommendations': risk_level['recommendations'],
                'has_post_surgery': hasattr(form, 'post_surgery_form'),
                'fecha_reporte': form.fecha_reporte
            }
            
            risk_assessments.append(assessment)
            
            # Generate alerts for high-risk patients
            if risk_score >= 70:  # Critical risk
                alert = generate_critical_alert(form, risk_score, risk_factors)
                high_risk_alerts.append(alert)
                critical_patients.append(assessment)
            elif risk_score >= 40:  # High risk
                alert = generate_high_risk_alert(form, risk_score, risk_factors)
                high_risk_alerts.append(alert)
                
        except Exception as e:
            print(f"Error calculating risk for {form.folio_hospitalizacion}: {e}")
            continue
    
    # Risk Distribution
    risk_distribution = {'low': 0, 'moderate': 0, 'high': 0, 'critical': 0}
    for assessment in risk_assessments:
        score = assessment['risk_score']
        if score >= 80:
            risk_distribution['critical'] += 1
        elif score >= 60:
            risk_distribution['high'] += 1
        elif score >= 40:
            risk_distribution['moderate'] += 1
        else:
            risk_distribution['low'] += 1
    
    # ASA Score Distribution
    asa_distribution = list(presurgery_forms
        .values('estado_fisico_asa')
        .annotate(count=Count('estado_fisico_asa'))
        .order_by('estado_fisico_asa'))
    
    # Mallampati Distribution
    mallampati_distribution = list(presurgery_forms
        .values('mallampati')
        .annotate(count=Count('mallampati'))
        .order_by('mallampati'))
    
    # Monthly Surgery Trends
    monthly_surgeries = list(presurgery_forms
        .annotate(month=TruncMonth('fecha_reporte'))
        .values('month')
        .annotate(count=Count('folio_hospitalizacion'))
        .order_by('month'))
    
    # BMI Distribution and Analysis
    bmi_stats = calculate_bmi_distribution(presurgery_forms)
    
    # Age Distribution
    age_distribution = calculate_age_distribution(user_patients)
    
    # Complications Analysis
    complications_data = analyze_complications(postsurgery_forms)
    
    # Intubation Success Analysis
    intubation_data = analyze_intubation_outcomes(postsurgery_forms)
    
    # Recent Activity Feed
    recent_activity = get_recent_activity(request.user, limit=10)
    
    # Performance Metrics
    performance_metrics = calculate_performance_metrics(
        presurgery_forms, postsurgery_forms, risk_assessments
    )
    
    # Upcoming Scheduled Procedures (if fecha_reporte is in future)
    upcoming_procedures = presurgery_forms.filter(
        fecha_reporte__gte=datetime.now().date()
    ).order_by('fecha_reporte')[:5]
    
    context = {
        # Basic Stats
        'total_patients': total_patients,
        'total_surgeries': total_surgeries,
        'completed_surgeries': completed_surgeries,
        'pending_surgeries': pending_surgeries,
        'patients_this_month': patients_this_month,
        
        # Risk Assessment Data
        'risk_assessments': risk_assessments[:10],  # Latest 10
        'high_risk_alerts': high_risk_alerts,
        'critical_patients': critical_patients[:5],  # Top 5 critical
        'risk_distribution': json.dumps(risk_distribution),
        'total_risk_assessments': len(risk_assessments),
        
        # Chart Data
        'asa_distribution': json.dumps(asa_distribution),
        'mallampati_data': json.dumps(mallampati_distribution),
        'monthly_surgeries': json.dumps([{
            'month': d['month'].strftime('%Y-%m-%d') if d['month'] else None,
            'count': d['count']
        } for d in monthly_surgeries]),
        'bmi_distribution': json.dumps(bmi_stats['distribution']),
        'age_distribution': json.dumps(age_distribution),
        
        # Analysis Data
        'complications_data': complications_data,
        'intubation_data': intubation_data,
        'performance_metrics': performance_metrics,
        'bmi_stats': bmi_stats,
        
        # Activity Data
        'recent_activity': recent_activity,
        'upcoming_procedures': upcoming_procedures,
        
        # Alert Counts
        'total_alerts': len(high_risk_alerts),
        'critical_alerts': len([a for a in high_risk_alerts if a['severity'] == 'CRITICAL']),
        'high_alerts': len([a for a in high_risk_alerts if a['severity'] == 'HIGH']),
    }
    
    return render(request, 'dashboard.html', context)

def generate_critical_alert(form, risk_score, risk_factors):
    """Generate critical risk alert"""
    return {
        'id': f"alert_{form.folio_hospitalizacion}",
        'severity': 'CRITICAL',
        'title': 'RIESGO CRÍTICO DE VÍA AÉREA',
        'patient': f"{form.nombres} {form.apellidos}",
        'folio': form.folio_hospitalizacion,
        'risk_score': risk_score,
        'message': f"Paciente con riesgo crítico ({risk_score}%) - Intervención inmediata requerida",
        'factors': risk_factors,
        'recommendations': [
            "🚨 Considerar intubación con paciente despierto",
            "🏥 Anestesiólogo senior requerido",
            "🛠️ Carro de vía aérea difícil disponible",
            "📋 Plan de vía aérea quirúrgica definido",
            "👥 Equipo de apoyo adicional"
        ],
        'timestamp': timezone.now(),
        'priority': 1,
        'color': '#dc3545',
        'icon': 'fas fa-exclamation-triangle'
    }

def generate_high_risk_alert(form, risk_score, risk_factors):
    """Generate high risk alert"""
    return {
        'id': f"alert_{form.folio_hospitalizacion}",
        'severity': 'HIGH',
        'title': 'RIESGO ALTO DE VÍA AÉREA',
        'patient': f"{form.nombres} {form.apellidos}",
        'folio': form.folio_hospitalizacion,
        'risk_score': risk_score,
        'message': f"Paciente con riesgo alto ({risk_score}%) - Preparación especial requerida",
        'factors': risk_factors,
        'recommendations': [
            "⚠️ Pre-oxigenación extendida recomendada",
            "📹 Videolaringoscopio disponible",
            "📋 Plan B de vía aérea definido",
            "🩺 Monitoreo estrecho",
            "🔄 Equipo de backup preparado"
        ],
        'timestamp': timezone.now(),
        'priority': 2,
        'color': '#ffc107',
        'icon': 'fas fa-exclamation-circle'
    }

def calculate_bmi_distribution(forms):
    """Calculate BMI distribution and statistics"""
    distribution = {
        'underweight': forms.filter(imc__lt=18.5).count(),
        'normal': forms.filter(imc__range=(18.5, 24.9)).count(),
        'overweight': forms.filter(imc__range=(25, 29.9)).count(),
        'obese_class_1': forms.filter(imc__range=(30, 34.9)).count(),
        'obese_class_2': forms.filter(imc__range=(35, 39.9)).count(),
        'obese_class_3': forms.filter(imc__gte=40).count()
    }
    
    # Calculate statistics
    bmi_values = forms.exclude(imc__isnull=True).values_list('imc', flat=True)
    if bmi_values:
        avg_bmi = sum(bmi_values) / len(bmi_values)
        high_bmi_patients = forms.filter(imc__gte=35).count()
        obesity_rate = (distribution['obese_class_1'] + distribution['obese_class_2'] + 
                       distribution['obese_class_3']) / max(len(bmi_values), 1) * 100
    else:
        avg_bmi = 0
        high_bmi_patients = 0
        obesity_rate = 0
    
    return {
        'distribution': distribution,
        'average_bmi': round(avg_bmi, 1),
        'high_bmi_patients': high_bmi_patients,
        'obesity_rate': round(obesity_rate, 1)
    }

def calculate_age_distribution(patients):
    """Calculate age distribution"""
    distribution = {
        'pediatric': 0,    # 0-17
        'young_adult': 0,  # 18-39
        'middle_age': 0,   # 40-64
        'elderly': 0       # 65+
    }
    
    today = datetime.now().date()
    
    for patient in patients:
        age = (today - patient.fecha_nacimiento).days // 365
        if age < 18:
            distribution['pediatric'] += 1
        elif age < 40:
            distribution['young_adult'] += 1
        elif age < 65:
            distribution['middle_age'] += 1
        else:
            distribution['elderly'] += 1
    
    return distribution

def analyze_complications(postsurgery_forms):
    """Analyze complications and outcomes"""
    total_surgeries = postsurgery_forms.count()
    
    if total_surgeries == 0:
        return {
            'total_surgeries': 0,
            'with_complications': 0,
            'with_morbidity': 0,
            'with_mortality': 0,
            'complications_percentage': 0,
            'morbidity_percentage': 0,
            'mortality_percentage': 0,
            'common_complications': []
        }
    
    with_complications = postsurgery_forms.filter(
        Q(complicaciones__isnull=False) & ~Q(complicaciones='')
    ).count()
    
    with_morbidity = postsurgery_forms.filter(morbilidad=True).count()
    with_mortality = postsurgery_forms.filter(mortalidad=True).count()
    
    # Analyze common complications
    complications_text = postsurgery_forms.exclude(
        Q(complicaciones__isnull=True) | Q(complicaciones='')
    ).values_list('complicaciones', flat=True)
    
    # Simple keyword analysis for common complications
    common_complications = analyze_complication_keywords(complications_text)
    
    return {
        'total_surgeries': total_surgeries,
        'with_complications': with_complications,
        'with_morbidity': with_morbidity,
        'with_mortality': with_mortality,
        'complications_percentage': round((with_complications / total_surgeries) * 100, 1),
        'morbidity_percentage': round((with_morbidity / total_surgeries) * 100, 1),
        'mortality_percentage': round((with_mortality / total_surgeries) * 100, 1),
        'common_complications': common_complications
    }

def analyze_complication_keywords(complications_text):
    """Analyze complications text for common keywords"""
    keywords = {
        'intubación difícil': 0,
        'broncoaspiración': 0,
        'hipoxemia': 0,
        'laringoespasmo': 0,
        'traumatismo dental': 0,
        'esofágica': 0,
        'neumotórax': 0
    }
    
    for text in complications_text:
        text_lower = text.lower()
        for keyword in keywords:
            if keyword in text_lower:
                keywords[keyword] += 1
    
    # Return top 5 complications
    sorted_complications = sorted(keywords.items(), key=lambda x: x[1], reverse=True)
    return [{'complication': k, 'count': v} for k, v in sorted_complications[:5] if v > 0]

def analyze_intubation_outcomes(postsurgery_forms):
    """Analyze intubation success rates and attempts"""
    total_intubations = postsurgery_forms.count()
    
    if total_intubations == 0:
        return {
            'total_attempts': 0,
            'first_attempt': 0,
            'multiple_attempts': 0,
            'success_rate': 0,
            'average_attempts': 0,
            'difficult_intubations': 0
        }
    
    first_attempt = postsurgery_forms.filter(numero_intentos=1).count()
    multiple_attempts = postsurgery_forms.filter(numero_intentos__gt=1).count()
    difficult_intubations = postsurgery_forms.filter(numero_intentos__gte=3).count()
    
    # Calculate average attempts
    attempts_data = postsurgery_forms.exclude(
        numero_intentos__isnull=True
    ).values_list('numero_intentos', flat=True)
    
    average_attempts = sum(attempts_data) / len(attempts_data) if attempts_data else 0
    
    return {
        'total_attempts': total_intubations,
        'first_attempt': first_attempt,
        'multiple_attempts': multiple_attempts,
        'success_rate': round((first_attempt / total_intubations) * 100, 1),
        'average_attempts': round(average_attempts, 1),
        'difficult_intubations': difficult_intubations
    }

def get_recent_activity(user, limit=10):
    """Get recent activity for the user"""
    activities = []
    
    # Recent patients
    recent_patients = Patient.objects.filter(
        medico=user, activo=True
    ).order_by('-fecha_registro')[:limit//2]
    
    for patient in recent_patients:
        activities.append({
            'type': 'patient_created',
            'title': 'Nuevo paciente registrado',
            'description': f"{patient.nombres} {patient.apellidos}",
            'timestamp': patient.fecha_registro,
            'icon': 'fas fa-user-plus',
            'color': 'success'
        })
    
    # Recent forms
    recent_forms = PreSurgeryForm.objects.filter(
        medico=user.get_full_name()
    ).order_by('-fecha_reporte')[:limit//2]
    
    for form in recent_forms:
        activities.append({
            'type': 'form_created',
            'title': 'Formulario pre-quirúrgico',
            'description': f"{form.nombres} {form.apellidos} - {form.fecha_reporte}",
            'timestamp': timezone.make_aware(datetime.combine(form.fecha_reporte, datetime.min.time())),
            'icon': 'fas fa-file-medical',
            'color': 'info'
        })
    
    # Sort by timestamp and return limited results
    activities.sort(key=lambda x: x['timestamp'], reverse=True)
    return activities[:limit]

def calculate_performance_metrics(presurgery_forms, postsurgery_forms, risk_assessments):
    """Calculate performance metrics for the doctor"""
    total_cases = len(risk_assessments)
    
    if total_cases == 0:
        return {
            'total_cases': 0,
            'average_risk_score': 0,
            'high_risk_cases': 0,
            'successful_outcomes': 0,
            'improvement_suggestions': []
        }
    
    # Calculate average risk score
    total_risk = sum(assessment['risk_score'] for assessment in risk_assessments)
    average_risk_score = total_risk / total_cases
    
    # Count high risk cases
    high_risk_cases = len([a for a in risk_assessments if a['risk_score'] >= 60])
    
    # Calculate successful outcomes (completed surgeries without major complications)
    successful_outcomes = postsurgery_forms.filter(
        Q(complicaciones__isnull=True) | Q(complicaciones=''),
        morbilidad=False,
        mortalidad=False
    ).count()
    
    # Generate improvement suggestions
    improvement_suggestions = generate_improvement_suggestions(
        risk_assessments, postsurgery_forms
    )
    
    return {
        'total_cases': total_cases,
        'average_risk_score': round(average_risk_score, 1),
        'high_risk_cases': high_risk_cases,
        'successful_outcomes': successful_outcomes,
        'success_rate': round((successful_outcomes / max(postsurgery_forms.count(), 1)) * 100, 1),
        'improvement_suggestions': improvement_suggestions
    }

def generate_improvement_suggestions(risk_assessments, postsurgery_forms):
    """Generate personalized improvement suggestions"""
    suggestions = []
    
    # Analyze patterns
    high_mallampati_cases = len([a for a in risk_assessments 
                                if hasattr(a['form'], 'mallampati') and a['form'].mallampati >= 3])
    
    high_asa_cases = len([a for a in risk_assessments 
                         if hasattr(a['form'], 'estado_fisico_asa') and a['form'].estado_fisico_asa >= 4])
    
    multiple_attempts = postsurgery_forms.filter(numero_intentos__gt=1).count()
    
    # Generate suggestions based on patterns
    if high_mallampati_cases > len(risk_assessments) * 0.3:
        suggestions.append({
            'title': 'Evaluación de Vía Aérea',
            'description': 'Considere usar videolaringoscopía más frecuentemente en casos Mallampati III-IV',
            'priority': 'high'
        })
    
    if high_asa_cases > len(risk_assessments) * 0.2:
        suggestions.append({
            'title': 'Pacientes ASA Alto Riesgo',
            'description': 'Implemente protocolos de pre-oxigenación extendida para pacientes ASA IV-V',
            'priority': 'medium'
        })
    
    if multiple_attempts > postsurgery_forms.count() * 0.15:
        suggestions.append({
            'title': 'Tasa de Intubación',
            'description': 'Revise técnicas de intubación - tasa de múltiples intentos superior al promedio',
            'priority': 'high'
        })
    
    return suggestions[:3]  # Return top 3 suggestions

@login_required
def get_dashboard_stats(request):
    """API endpoint for real-time dashboard statistics"""
    try:
        date_range = request.GET.get('date_range', '30')
        start_date = datetime.now() - timedelta(days=int(date_range))
        
        # Get filtered data
        user_patients = Patient.objects.filter(medico=request.user, activo=True)
        presurgery_forms = PreSurgeryForm.objects.filter(
            medico=request.user.get_full_name(),
            fecha_reporte__gte=start_date.date()
        )
        postsurgery_forms = PostDuringSurgeryForm.objects.filter(
            folio_hospitalizacion__in=presurgery_forms.values('folio_hospitalizacion')
        )

        # Calculate real-time risk assessments
        current_alerts = []
        risk_summary = {'low': 0, 'moderate': 0, 'high': 0, 'critical': 0}
        
        for form in presurgery_forms:
            try:
                risk_score, risk_factors = RiskCalculator.calculate_airway_risk(form)
                
                if risk_score >= 80:
                    risk_summary['critical'] += 1
                    current_alerts.append({
                        'severity': 'CRITICAL',
                        'patient': f"{form.nombres} {form.apellidos}",
                        'risk_score': risk_score,
                        'factors': risk_factors[:3]  # Top 3 factors
                    })
                elif risk_score >= 60:
                    risk_summary['high'] += 1
                elif risk_score >= 40:
                    risk_summary['moderate'] += 1
                else:
                    risk_summary['low'] += 1
                    
            except Exception:
                continue

        stats = {
            'summary': {
                'total_patients': user_patients.count(),
                'active_patients': user_patients.filter(
                    fecha_registro__gte=start_date
                ).count(),
                'surgeries_completed': postsurgery_forms.count(),
                'pending_surgeries': presurgery_forms.filter(
                    folio_hospitalizacion__in=presurgery_forms.exclude(
                        folio_hospitalizacion__in=postsurgery_forms.values(
                            'folio_hospitalizacion__folio_hospitalizacion'
                        )
                    ).values('folio_hospitalizacion')
                ).count()
            },
            
            'risk_summary': risk_summary,
            'current_alerts': current_alerts[:5],  # Top 5 alerts
            
            'asa_distribution': [
                {
                    'estado_fisico_asa': item['estado_fisico_asa'],
                    'count': item['count']
                }
                for item in presurgery_forms.values('estado_fisico_asa')
                .annotate(count=Count('estado_fisico_asa'))
                .order_by('estado_fisico_asa')
            ],
            
            'recent_surgeries': [
                {
                    'week': item['week'].strftime('%Y-%m-%d') if item['week'] else None,
                    'count': item['count']
                }
                for item in presurgery_forms
                .annotate(week=TruncWeek('fecha_reporte'))
                .values('week')
                .annotate(count=Count('folio_hospitalizacion'))
                .order_by('week')
            ],
            
            'complications_summary': {
                'total_completed': postsurgery_forms.count(),
                'with_complications': postsurgery_forms.filter(
                    Q(complicaciones__isnull=False) & ~Q(complicaciones='')
                ).count(),
                'morbidity_cases': postsurgery_forms.filter(morbilidad=True).count(),
                'mortality_cases': postsurgery_forms.filter(mortalidad=True).count()
            },
            
            'intubation_metrics': {
                'first_attempt_success': postsurgery_forms.filter(numero_intentos=1).count(),
                'multiple_attempts': postsurgery_forms.filter(numero_intentos__gt=1).count(),
                'difficult_cases': postsurgery_forms.filter(numero_intentos__gte=3).count()
            }
        }
        
        return JsonResponse(stats)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_patient_alerts(request):
    """API endpoint for getting current patient alerts"""
    try:
        presurgery_forms = PreSurgeryForm.objects.filter(
            medico=request.user.get_full_name()
        )
        
        alerts = []
        
        for form in presurgery_forms:
            try:
                risk_score, risk_factors = RiskCalculator.calculate_airway_risk(form)
                
                if risk_score >= 70:  # High risk threshold
                    alert = {
                        'id': f"alert_{form.folio_hospitalizacion}",
                        'severity': 'CRITICAL' if risk_score >= 80 else 'HIGH',
                        'patient_name': f"{form.nombres} {form.apellidos}",
                        'folio': form.folio_hospitalizacion,
                        'risk_score': risk_score,
                        'risk_factors': risk_factors,
                        'timestamp': timezone.now().isoformat(),
                        'has_post_surgery': hasattr(form, 'post_surgery_form')
                    }
                    alerts.append(alert)
                    
            except Exception as e:
                continue
        
        # Sort by risk score (highest first)
        alerts.sort(key=lambda x: x['risk_score'], reverse=True)
        
        return JsonResponse({
            'alerts': alerts,
            'total_alerts': len(alerts),
            'critical_count': len([a for a in alerts if a['severity'] == 'CRITICAL']),
            'high_count': len([a for a in alerts if a['severity'] == 'HIGH'])
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required 
def dismiss_alert(request, alert_id):
    """API endpoint to dismiss an alert"""
    if request.method == 'POST':
        try:
            # In a real implementation, you'd store dismissed alerts in the database
            # For now, we'll just return success
            return JsonResponse({'status': 'success', 'message': 'Alert dismissed'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def export_dashboard(request):
    """Export dashboard data to Excel"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Export"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B4570", end_color="2B4570", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Patient Summary
        ws.append(["Resumen de Pacientes"])
        ws.append(["Fecha de Exportación", datetime.now().strftime("%Y-%m-%d %H:%M")])
        ws.append([])

        # Style headers
        for cell in ws[1:1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Add statistics sections
        sections = [
            ("Información General", get_general_stats(request.user)),
            ("Estadísticas ASA", get_asa_stats(request.user)),
            ("Complicaciones", get_complication_stats(request.user)),
            ("Métricas de Vía Aérea", get_airway_stats(request.user))
        ]

        row = 4
        for section_title, data in sections:
            ws.append([section_title])
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            
            row += 1
            for key, value in data.items():
                ws.append([key, value])
            
            row += 2

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return file
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f'dashboard_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Helper functions for export
def get_general_stats(user):
    return {
        "Total Pacientes": Patient.objects.filter(medico=user, activo=True).count(),
        "Cirugías Este Mes": PreSurgeryForm.objects.filter(
            medico=user.get_full_name(),
            fecha_reporte__month=datetime.now().month
        ).count(),
        "Pacientes de Alto Riesgo": PreSurgeryForm.objects.filter(
            medico=user.get_full_name(),
            estado_fisico_asa__gte=4
        ).count()
    }

def get_asa_stats(user):
    return dict(PreSurgeryForm.objects.filter(
        medico=user.get_full_name()
    ).values('estado_fisico_asa').annotate(
        count=Count('estado_fisico_asa')
    ).values_list('estado_fisico_asa', 'count'))

def get_complication_stats(user):
    postsurgery_forms = PostDuringSurgeryForm.objects.filter(
        nombre_anestesiologo=user.get_full_name()
    )
    total = postsurgery_forms.count() or 1
    complications = postsurgery_forms.filter(
        complicaciones__isnull=False
    ).exclude(complicaciones='').count()
    
    return {
        "Total Cirugías": total,
        "Cirugías con Complicaciones": complications,
        "Tasa de Complicaciones": f"{(complications/total)*100:.2f}%"
    }

def get_airway_stats(user):
    presurgery_forms = PreSurgeryForm.objects.filter(
        medico=user.get_full_name()
    )
    return {
        "Vía Aérea Difícil": presurgery_forms.filter(
            Q(mallampati__gte=3) |
            Q(patil_aldrete__gte=3)
        ).count(),
        "Mallampati Promedio": presurgery_forms.aggregate(
            Avg('mallampati')
        )['mallampati__avg']
    }

def calculate_average_age(patients):
    total_age = 0
    count = 0
    for patient in patients:
        try:
            age = (datetime.now().date() - patient.fecha_nacimiento).days // 365
            total_age += age
            count += 1
        except:
            continue
    return round(total_age / max(count, 1), 1)


@login_required
def presurgery_detail(request, pk):
    # Get the pre-surgery form
    form = get_object_or_404(PreSurgeryForm, folio_hospitalizacion=pk)
    
    # Get the related patient
    patient = get_object_or_404(Patient, folio_hospitalizacion=pk.replace('PRE-', ''))
    
    # Check if user has permission to view this form
    if patient.medico != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para ver este formulario.')
        return redirect('patient-list')
    
    context = {
        'form': form,
        'patient': patient
    }
    
    return render(request, 'presurgery_detail.html', context)


# Add these error handlers to views.py
def custom_404(request, exception):
    return render(request, 'errors/404.html', status=404)

def custom_500(request):
    return render(request, 'errors/500.html', status=500)